
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from contextlib import redirect_stdout 
import xarray as xr
import pandas as pd


def parse_brick_F2D_CSV_data(brick_ds, file_name):
    # Read df
    df = pd.read_csv(file_name, sep='\t') 

    # Build vairable values for each dims
    dims = []

    # First dim
    brick_dims = brick_ds['dimensions'] 
    if len(brick_dims) > 0:
        dim = {
            'size':  0,
            'dim_vars': []
        }
        dims.append(dim)
        for i,v in enumerate(brick_dims[0]['variables']):
            vals = list(df[df.columns[i]].values)
            dim['size'] = len(vals)
            dim['dim_vars'].append({
                'values': vals
            })

    # Second dim
    if len(brick_dims) > 1:
        dim = {
            'size':  0,
            'dim_vars': []
        }
        dims.append(dim)
        offset = len(brick_dims[0]['variables'])
        vals = list(df.columns[offset:].values)
        dim['size'] = len(vals)
        dim['dim_vars'].append({
                'values': vals
            })

    data_vars = []
    offset = len(dims[0]['dim_vars'])
    vals = df[df.columns[offset:].values].values
    data_vars.append({
                'values': vals
            })

    _validate(dims, data_vars)    
    return {
        'dims': dims,
        'data_vars': data_vars
    }

def _validate(dims, data_vars):  
    dim_sizes = []

    # Validate dimension variables
    for dim in dims:
        dim_size = None
        for dim_var in dim['dim_vars']:
            size = len(dim_var['values'])
            if dim_size is None:
                dim_size = size
            else:
                if dim_size != size:
                    raise ValueError('Size of two vars from the same dimension is different: %s and %s' % (dim_size, size))  
        dim_sizes.append(dim_size)

    # Validate data vars
    for data_var in data_vars:
        data_shape = list(np.array(data_var['values']).shape)

        if len(data_shape) != len(dim_sizes):
            raise ValueError('The dimensionality of data (%s) is different from the context (%s)' % (len(data_shape), len(dim_sizes)))

        for dim_index, dim_size in enumerate(dim_sizes):
            if dim_size != data_shape[dim_index]:
                raise ValueError('The size of dimension for data var (%s) and context (%s) ' % (dim_size, data_shape[dim_index]))

def _validate_interlaced_dims(sheet, n_data_vars, start_row, end_row, start_col, end_col):
    concat_dim_cols = []
    for ci, col in enumerate(sheet.iter_cols(min_row=start_row,
                                             max_row=end_row,
                                             min_col=start_col,
                                             max_col=end_col)):
        concat = ''
        for cell in col:
            concat += '_' + str(cell.value)
        if concat in concat_dim_cols and ci % n_data_vars == 0:
            raise ValueError('Incorrect Variable Format, you have repeating interlaced variable columns. Did you mean to add a replicate series dimension?')
        else:
            concat_dim_cols.append(concat)

def _get_1d_column_data(sheet, ri, ci):
    data = []
    while True:
        value = sheet.cell(row=ri, column=ci ).value
        if value is None:
            break
        if value in ['-', 'n/a', 'N/A']:
            value = None
        data.append(value)
        ri += 1
    return data

def _get_1d_row_data(sheet, ri, ci):
    data = []
    while True:
        value = sheet.cell(row=ri, column=ci ).value
        if value is None:
            break
        if value in ['-', 'n/a', 'N/A']:
            value = None
        data.append(value)
        ci += 1
    return data

def _get_2d_data(sheet, ri_start, ci_start, r_len, c_len):
    data = []
    for ri in range(ri_start, ri_start + r_len):
        row = []
        for ci in range(ci_start, ci_start + c_len):
            value = sheet.cell(row=ri, column=ci).value
            row.append(value)
        data.append(row)
    return data


def _get_contexted_var_name(dvar):
    var_name = dvar['type']['text']
    if 'units' in dvar and dvar['units'] and dvar['units']['text']:
        var_name += ' (%s)' % dvar['units']['text']

    items = []
    if 'context' in dvar:
        for ce in dvar['context']:
            item = '%s=%s' % (ce['type']['text'], ce['value']['text'])
            if 'units' in ce and ce['units'] and ce['units']['text']:
                item += ' (%s)' % ce['units']['text']
            items.append(item)

    return var_name if len(items) == 0 else '%s:%s' % (var_name, '; '.join(items))

def generate_brick_template(brick_ds, file_name, tab_data=False):

    """
    generates excel template for brick with any number of dimensions.
    generates multiple tabs within a sheet. if tab_data is True, it will tab across individual data vars,
    otherwise will tab across the highest dimension variable
    """

    TITLE_FONT = Font(name='Calibri', size=14)
    INSTRUCTIONS_FONT = Font(name='Calibri', size=10, bold=True, italic=True)
    DIM1_FILL = PatternFill('solid', fgColor='FFFFDD')
    DIM2_FILL = PatternFill('solid', fgColor='DDFFDD')
    DIM3_FILL = PatternFill('solid', fgColor='DDDDFF')
    DATA_FILL = PatternFill('solid', fgColor='FFDDDD')
    COLUMN_WIDTH = 18
    dim1_row_counts = 10
    dim2_row_counts = 10
    dim3_row_counts = 10
    FORMAT_NAME = 'FND_TAB_DATA' if tab_data else 'FND_TAB_DIM'

    book = Workbook()
    sheets = [book.active]

    book.active.title = brick_ds['dataValues'][0]['type']['text'] if tab_data else 'Value 1'

    template_name = brick_ds['template_type']
    data_vars = brick_ds['dataValues']
    brick_dims = brick_ds['dimensions']
    data_type = brick_ds['type']['text']

    if tab_data:
        for data_value in brick_ds['dataValues'][1:]:
            sheets.append(book.create_sheet(data_value['type']['text']))
      #  for _ in range(len(brick_ds['dataValues']) - 1):
      #      sheets.append(book.create_sheet())
    else:
        for i in range(dim3_row_counts - 1):
            sheets.append(book.create_sheet('Value ' + str(i + 2)))

    for si, sheet in enumerate(sheets):

        if tab_data:
            data_var = _get_contexted_var_name(data_vars[si])
        else:
            data_var = _get_contexted_var_name(data_vars[0])

        dims = [brick_dims[0]['type']['text']]
        interlace_dims = [dim['type']['text'] for dim in brick_dims[1:]]
    
        dim_vars = []
        for dim in brick_dims:
            _dim_vars = []
            for dim_var in dim['variables']:
                var_name = _get_contexted_var_name(dim_var)
                _dim_vars.append('%s:%s' % (dim['type']['text'], var_name))
            dim_vars.append(_dim_vars)

        # create dynamic dimension color fills
        pattern_fills = []
        pattern_fill_range = 3 if len(brick_dims) < 3 else len(brick_dims)
        color = 221
        for i in range(pattern_fill_range):
            hex = f'{color:X}'
            if i % 6 == 0:
                if i > 0:
                    color -= 16
                pattern = 'FFFF' + hex
            elif i % 6 == 1:
                pattern = hex + 'FF' + hex
            elif i % 6 == 2:
                pattern = hex + hex + 'FF'
            elif i % 6 == 3:
                pattern = 'FF' + hex + 'FF'
            elif i % 6 == 4:
                pattern = hex + 'FFFF'
            elif i % 6 == 5:
                pattern = hex + hex + hex
            pattern_fills.append(PatternFill('solid', fgColor=pattern))


        rows = [
            ['Automatically generated template'],
            ['Name', template_name],
            ['Type', data_type]
        ]

        for i, dim in enumerate(dims):
            rows.append(['Dimension ' + str(i + 1), dim])

        for i, dim in enumerate(interlace_dims):
            rows.append(['Dimension ' + str(i + len(dims) + 1), dim])

        if tab_data:
            rows.append(['Data Value', data_var])
        else:
            for di, dv in enumerate(data_vars):
                rows.append(['Data Value ' + str(di + 1), _get_contexted_var_name(dv)])

        rows.extend([
            [],
            ['Instructions: fill in all colored parts of the spreadsheet below with your data and metadata'],
            ['@Format:%s' % FORMAT_NAME],
        ])

        if not tab_data:
           rows.extend([['Fill out ' + str(brick_dims[-1]['type']['text']) + ' variables in the following highlighted area']])
        else:
           rows.extend([[]])

        #generate label for tabbed item (above spreadsheet)

        if not tab_data:
            last_dim_row_labels = []
            last_dim_rows = []

            for dim in dim_vars[-1]:
                last_dim_row_labels.append(dim)
                last_dim_rows.append('')

            rows.append(last_dim_row_labels)
            rows.append(last_dim_rows)

        # add interlaced 2nd+ dimensions rows
        interlace_rows = dim_vars[1:] if tab_data else dim_vars[1:-1]
        for interlace in interlace_rows:
            for dim_var in interlace:
                row = ['' for _ in dim_vars[0]]
                row.append(dim_var)
                rows.append(row)

        # add 1st dimension columns
        row = []
        for dim1_var in dim_vars[0]:
            row.append(dim1_var)
        rows.append(row)

        for row in rows:
            sheet.append(row)

        sheet['A1'].font = TITLE_FONT

        for i, pattern_fill in enumerate(pattern_fills):
            sheet.cell(row=i + 4, column=1).fill = pattern_fill
            sheet.cell(row=i + 4, column=2).fill = pattern_fill

        if not tab_data:
            color = 221
            min_row = len(brick_dims) + 4
            data_pattern_fills = []
            for row in sheet.iter_rows(min_row=min_row, max_col=2, max_row=len(data_vars) + min_row - 1):
                color_hex = 'FF' + f'{color:X}' + f'{color:X}'
                pattern_fill = PatternFill('solid', fgColor=color_hex)
                data_pattern_fills.append(pattern_fill)
                for cell in row:
                    cell.fill = pattern_fill
                color -= 16
        else:
            data_var_idx = len(brick_dims) + 4
            sheet[data_var_idx][0].fill = DATA_FILL
            sheet[data_var_idx][1].fill = DATA_FILL

        # add formula for 1st dimension columns
        if si > 0:
            for row in sheet.iter_rows(min_row=start_row, max_col=len(dim_vars[0]) + 1, max_row=1000,min_col=1):
                for cell in row:
                    master_sheet = brick_ds['dataValues'][0]['type']['text'] if tab_data else 'Value 1'
                    coord = cell.coordinate
                    cell.value = f'=IF(ISBLANK(INDIRECT("\'{master_sheet}\'!{coord}")),"",INDIRECT("\'{master_sheet}\'!{coord}"))'

        # TODO: dynamically calculate where instructions are 

        # Color dim 1
        start_row = len(rows) + 1
        start_column = 1
        for ri in range(start_row, start_row + dim1_row_counts):
            for ci, _ in enumerate(dim_vars[0]):
                sheet.cell(row=ri, column=ci+start_column).fill = DIM1_FILL

        # Color tab label
        interlace_height = len([j for sub in interlace_rows for j in sub])
        start_row = len(rows) - interlace_height - 1
        if not tab_data:
            for di, dv in enumerate(dim_vars[-1]):
                sheet.cell(row=start_row, column=di+1).fill = pattern_fills[-1]
 
        # Color interlaced dimensions
        start_row = len(rows) - interlace_height
        for ri, interlace_row in enumerate(interlace_rows):
            for row in sheet.iter_rows(min_col=len(dim_vars[0]) + 2,
                                       max_col=dim2_row_counts + len(dim_vars[0]) + 1,
                                       min_row=start_row,
                                       max_row=start_row + len(interlace_row) - 1):
                for cell in row:
                    cell.fill = pattern_fills[ri + 1]
            start_row += len(interlace_row)

        # add formula for interlaced dimension rows
        if si > 0:
            start_row= len(rows) - interlace_height
            for row in sheet.iter_rows(min_col=len(dim_vars[0]) + 2, max_col=1000, min_row=start_row, max_row=start_row + interlace_height - 1):
                for cell in row:
                    master_sheet = brick_ds['dataValues'][0]['type']['text'] if tab_data else 'Value 1'
                    coord = cell.coordinate
                    cell.value = f'=IF(ISBLANK(INDIRECT("\'{master_sheet}\'!{coord}")),"",INDIRECT("\'{master_sheet}\'!{coord}"))'

        # Color data
        start_row = len(rows) + 1
        start_column = len(dim_vars[0]) + 2
        if tab_data:
            for ri in range(start_row, start_row + dim1_row_counts):
                for ci in range(start_column, start_column + dim2_row_counts):
                    sheet.cell(row=ri, column=ci).fill = DATA_FILL
        else:
            for ri in range(start_row, start_row + dim1_row_counts):
                for ci in range(start_column, start_column + dim2_row_counts):
                    idx = start_column - ci
                    data_length = len(data_pattern_fills)
                    # sheet.cell(row=ri, column=ci).fill = data_pattern_fills[idx % data_length]
                    sheet.cell(row=ri, column=ci).fill = data_pattern_fills[data_length - ((idx - 1) % data_length) - 1]

        # label data fields
        if not tab_data and len(data_vars) > 1:
            label_row = len(rows)
            start_column = len(dim_vars[0]) + 2
            for ci in range(start_column, start_column + dim2_row_counts):
                idx = start_column - ci
                data_length = len(data_vars)
                _data_var = data_vars[data_length - ((idx - 1) % data_length) - 1]
                sheet.cell(row=label_row, column=ci).value = _get_contexted_var_name(_data_var)

        for i in range(1, len(dim_vars[0]) + 2):
            c = get_column_letter(i)
            sheet.column_dimensions[c].width = COLUMN_WIDTH


        # create example data
        sheet['I2'].value = 'Example: fill out dimension 1 on the side, then fill out the remaining dimensions such that dimension 2 permutes with dimension 3, etc.'
        if tab_data or len(data_vars) == 1:
            for ri, row in enumerate(sheet.iter_rows(min_row=3, max_row=4, min_col=11, max_col=15)):
                row[0].value = 'Dim ' + str(ri + 2)
                for ci, cell in enumerate(row[1:]):
                    cell.fill = pattern_fills[ri + 1]
                    if ri == 0:
                        cell.value = 'Value 1' if ci < 2 else 'Value 2'
                    else: 
                        cell.value = 'Value ' + str((ci % 2) + 1)
            sheet['J5'].value = 'Dim 1'
            for ri, row in enumerate(sheet.iter_rows(min_row=6, max_row=10, min_col=10, max_col=15)):
                for ci, cell in enumerate(row):
                    if ci == 0:
                        cell.fill = pattern_fills[0]
                        cell.value = 'Value' + str(ri + 1)
                    elif ci > 1:
                        cell.fill = DATA_FILL
                        cell.value = 'Data'
        else:
            var_length = len(dim_vars)
            data_length = len(data_vars)

            for ri, row in enumerate(sheet.iter_rows(min_row=3, max_row=4, min_col=11, max_col=21)):
                row[0].value = 'Dim ' + str(ri + 2)
                if ri == 0:
                    idx = 1
                    for ci, cell in enumerate(row[1:]):
                        if ci > 0 and ci % (data_length * 2) == 0:
                            idx += 1
                        cell.value = 'Var ' + str(idx)
                        cell.fill = pattern_fills[ri + 1]
                else:
                    idx = 1
                    for ci, cell in enumerate(row[1:]):
                        if ci > 0 and ci % data_length == 0:
                            idx = 1 if idx + 1 > 2 else idx + 1
                        cell.value = 'Var ' + str(idx)
                        cell.fill = pattern_fills[ri + 1]
            sheet['J5'].value = 'Dim 1'
            for ri, row in enumerate(sheet.iter_rows(min_row=6, max_row=10, min_col=10, max_col=21)):
                for ci, cell in enumerate(row):
                    if ci == 0:
                        cell.fill = pattern_fills[0]
                        cell.value = 'Var ' + str(ri + 1)
                    elif ci > 1:
                        idx = ci - 2
                        cell.fill = data_pattern_fills[idx % data_length]
                        cell.value = 'Data'

        sheet['I2'].font = INSTRUCTIONS_FONT

    book.save(file_name)

def generate_brick_interlace_template(brick_ds, file_name):

    """
    generates template with all dim vars and data vars interlaced on one page (no tabs)
    """

    TITLE_FONT = Font(name='Calibri', size=14)
    INSTRUCTIONS_FONT = Font(name='Calibri', size=10, bold=True, italic=True)
    COLUMN_WIDTH = 18
    dim1_row_counts = 10
    dim2_row_counts = 10 if len(brick_ds['dimensions']) > 1 else len(brick_ds['dataValues'])
    DATA_FILL = PatternFill('solid', fgColor='FFDDDD')
    FORMAT_NAME = 'FND_INTERLACE'

    book = Workbook()
    sheet = book.active

    template_name = brick_ds['template_type']
    data_vars = brick_ds['dataValues']
    brick_dims = brick_ds['dimensions']
    data_type = brick_ds['type']['text']

    data_var = _get_contexted_var_name(data_vars[0])
    dims = [brick_dims[0]['type']['text']] # TODO: this might not be necessary
    interlace_dims = [dim['type']['text'] for dim in brick_dims[1:]]

    dim_vars = []
    for dim in brick_dims:
        _dim_vars = []
        for dim_var in dim['variables']:
            var_name = _get_contexted_var_name(dim_var)
            _dim_vars.append('%s:%s' % (dim['type']['text'], var_name))
        dim_vars.append(_dim_vars)

    # create dynamic dimension color fills 
    pattern_fills = []
    color = 221
    brick_dims_range = len(brick_dims) if len(brick_dims) > 3 else 3
    for i in range(brick_dims_range):
    #for i, _ in enumerate(brick_dims):
        hex = f'{color:02X}'
        if i % 6 == 0:
            if i > 0:
                color -= 16
            pattern = 'FFFF' + hex
        elif i % 6 == 1:
            pattern = hex + 'FF' + hex
        elif i % 6 == 2:
            pattern = hex + hex + 'FF'
        elif i % 6 == 3:
            pattern = 'FF' + hex + 'FF'
        elif i % 6 == 4:
            pattern = hex + 'FFFF'
        elif i % 6 == 5:
            patern = hex + hex + hex
        pattern_fills.append(PatternFill('solid', fgColor=pattern))

    rows = [
        ['Automatically generated template'],
        ['Name', template_name],
        ['Type', data_type]
    ]

    for i, dim in enumerate(dims):
        rows.append(['Dimension ' + str(i + 1), dim])

    for i, dim in enumerate(interlace_dims):
        rows.append(['Dimension ' + str(i + len(dims) + 1), dim])

    for di, dv in enumerate(data_vars):
        rows.append(['Data Value ' + str(di + 1), _get_contexted_var_name(dv)])

    rows.extend([
        [],
        ['Instructions: fill in all colored parts of the spreadsheet below with your data and metadata'],
        ['@Format %s' % FORMAT_NAME],
        []
    ])

    # add interlaced 2nd+ dimension rows
    interlace_rows = dim_vars[1:]
    for interlace in interlace_rows:
        for dim_var in interlace:
            row = ['' for _ in dim_vars[0]]
            row.append(dim_var)
            rows.append(row)

    # add first dimension columns
    row = []
    for dim1_var in dim_vars[0]:
        row.append(dim1_var)
    rows.append(row)

    for row in rows:
        sheet.append(row)

    sheet['A1'].font = TITLE_FONT

    for i in range(len(brick_dims)):
        sheet.cell(row=i + 4, column=1).fill = pattern_fills[i]
        sheet.cell(row=i + 4, column=2).fill = pattern_fills[i]

#    for i, pattern_fill in enumerate(pattern_fills):
#        sheet.cell(row=i + 4, column=1).fill = pattern_fill
#        sheet.cell(row=i + 4, column=2).fill = pattern_fill

    data_color = 221
    data_sub_color = 255
    min_row = len(brick_dims) + 4
    data_pattern_fills = []
    for row in sheet.iter_rows(min_row=min_row, max_col=2, max_row=len(data_vars) + min_row - 1):
        if data_color < 100:
            data_color = 221
            data_sub_color -= 15
        color_hex = f'{data_sub_color:02X}' + f'{data_color:02X}' + f'{data_color:02X}'
        pattern_fill = PatternFill('solid', fgColor=color_hex)
        data_pattern_fills.append(pattern_fill)
        for cell in row:
            cell.fill = pattern_fill
        data_color -= 16

    # color dim 1
    start_row = len(rows) + 1
    start_column = 1
    for ri in range(start_row, start_row + dim1_row_counts):
        for ci, _ in enumerate(dim_vars[0]):
            sheet.cell(row=ri, column=ci+start_column).fill = pattern_fills[0]

    # color interlaced dimensions
    interlace_height = len([j for sub in interlace_rows for j in sub])
    start_row = len(rows) - interlace_height
    for ri, interlace_row in enumerate(interlace_rows):
        for row in sheet.iter_rows(min_col=len(dim_vars[0]) + 2,
                                   max_col=dim2_row_counts + len(dim_vars[0]) + 1,
                                   min_row=start_row,
                                   max_row=start_row + len(interlace_row) - 1):
            for cell in row:
                cell.fill = pattern_fills[ri + 1]
        start_row += len(interlace_row)

    # color data
    start_row = len(rows) + 1
    start_column = len(dim_vars[0]) + 2
    for ri in range(start_row, start_row + dim1_row_counts):
        for ci in range(start_column, start_column + dim2_row_counts):
            idx = start_column - ci
            data_length = len(data_pattern_fills)
            sheet.cell(row=ri, column=ci).fill =  data_pattern_fills[data_length - ((idx - 1) % data_length) - 1]
    for i in range(1, len(dim_vars[0]) + 2):
        c = get_column_letter(i)
        sheet.column_dimensions[c].width = COLUMN_WIDTH

    # label data fields
    if len(data_vars) > 1:
        label_row = len(rows)
        start_column = len(dim_vars[0]) + 2
        for ci in range(start_column, start_column + dim2_row_counts):
            idx = start_column - ci
            data_length = len(data_vars)
            _data_var = data_vars[data_length - ((idx - 1) % data_length) - 1]
            sheet.cell(row=label_row, column=ci).value = _get_contexted_var_name(_data_var)

    # create example data
    sheet['I2'].value = 'Example: fill out dimension 1 on the side, then fill out the remaining dimensions such that dimension 2 permutes with dimension 3, etc.'
    if len(data_vars) == 1:
        for ri, row in enumerate(sheet.iter_rows(min_row=3, max_row=4, min_col=11, max_col=15)):
            row[0].value = 'Dim ' + str(ri + 2)
            for ci, cell in enumerate(row[1:]):
                cell.fill = pattern_fills[ri + 1]
                if ri == 0:
                    cell.value = 'Value 1' if ci < 2 else 'Value 2'
                else:
                    cell.value = 'Value ' + str((ci % 2) + 1)
        sheet['J5'].value = 'Dim 1'
        for ri, row in enumerate(sheet.iter_rows(min_row=6, max_row=10, min_col=10, max_col=15)):
            for ci, cell in enumerate(row):
                if ci == 0:
                    cell.fill = pattern_fills[0]
                    cell.value = 'Data'
                elif ci > 1:
                    cell.fill = DATA_FILL
                    cell.value = 'Data'
    else:
        var_length = len(dim_vars)
        data_length = len(data_vars)
        
        for ri, row in enumerate(sheet.iter_rows(min_row=3, max_row=4, min_col=11, max_col=21)):
            row[0].value = 'Dim ' + str(ri + 2)
            if ri == 0:
               idx = 1
               for ci, cell in enumerate(row[1:]):
                   if ci > 0 and ci % (data_length * 2) == 0:
                       idx += 1
                   cell.value = 'Var ' + str(idx)
                   cell.fill = pattern_fills[ri + 1]
            else:
                idx = 1
                for ci, cell in enumerate(row[1:]):
                    if ci > 0 and ci % data_length == 0:
                        idx = 1 if idx + 1 > 2 else idx + 1
                    cell.value = 'Var ' + str(idx)
                    cell.fill = pattern_fills[ri + 1]
        sheet['J5'].value = 'Dim 1'
        for ri, row in enumerate(sheet.iter_rows(min_row=6, max_row=10, min_col=10, max_col=21	)):
            for ci, cell in enumerate(row):
                if ci == 0:
                    cell.fill = pattern_fills[0]
                    cell.value = 'Var ' + str(ri + 1)
                elif ci > 1:
                    idx = ci - 2
                    cell.fill = data_pattern_fills[(idx % data_length)]
                    cell.value = 'Data'

    sheet['I2'].font = INSTRUCTIONS_FONT

    book.save(file_name)

def parse_brick_data(brick_ds, file_name):
    SEARCH_FORMAT_MAX_ROW = 100
    book = load_workbook(file_name)
    sheet = book.active
    template_formate = None
    for ri in range(1, SEARCH_FORMAT_MAX_ROW):
        value = sheet.cell(row=ri, column=1).value
        if value is not None and value.startswith('@Format'):
            template_format = value[len('@Format:'):]
            break

    if template_format is None:
        raise ValueError('Cannot identify the template format')

    if template_format == 'FND_INTERLACE':
        return parse_brick_interlace_data(brick_ds, sheet)
    elif template_format == 'FND_TAB_DIM':
        return parse_brick_tab_dim(brick_ds, book)
    elif template_format == 'FND_TAB_DATA':
        return parse_brick_tab_data(brick_ds, book)
    else:
        raise ValueError('Wrong data format: the expected format is FND_INTERLACE, FND_TAB_DIM, or FND_TAB_DATA; found %s' % template_format)

def parse_brick_interlace_data(brick_ds, sheet):

    dims_ds = brick_ds['dimensions']

    template_dims = [dim['type']['text'] for dim in dims_ds]
    template_dim_vars = []
    for di, dim in enumerate(dims_ds):
        template_dim_vars.append([])
        for dim_var in dim['variables']:
            var_name = _get_contexted_var_name(dim_var)
            template_dim_vars[-1].append('%s:%s' % (template_dims[di], var_name))

    dims = [{'size': 0, 'dim_vars': []} for _ in dims_ds]
    data_vars = []

    start_row = len(dims_ds) + len(brick_ds['dataValues']) + 8
    # Do dimension 1 variables
    interlace_dim_var_length = 0
    for dim in template_dim_vars[1:]:
        interlace_dim_var_length += len(dim)
    dim1_start_row = start_row + interlace_dim_var_length
    dim1_ci = 0
    for dim_var in template_dim_vars[0]:
        dim1_ci += 1
        header = sheet.cell(row=dim1_start_row, column=dim1_ci).value
        if header != dim_var:
            raise ValueError('Wrong format: expected dimension variable %s; found %s' % (dim_var, header))
        vals = _get_1d_column_data(sheet, dim1_start_row + 1, dim1_ci)
        dims[0]['size'] = len(vals)
        dims[0]['dim_vars'].append({
            'values': vals
        })
    # Do interlaced dimension variables
    dim_start_row = start_row
    dim_ci = len(template_dim_vars[0]) + 1
    for di, dim_vars in enumerate(template_dim_vars[1:]):
        idx = di + 1
        concat_vars = []
        for dim_var in dim_vars:
            header = sheet.cell(row=dim_start_row, column=dim_ci).value
            if header != dim_var:
                raise ValueError('Wrong format: expected dimension variable %s, found %s' % (dim_var, header))
            raw_vars = _get_1d_row_data(sheet, dim_start_row, dim_ci + 1)
            concat_vars.append(raw_vars)
            dim_start_row += 1
        if len(concat_vars) > 1:
            _concat_vars = []
            try:
                for i in range(len(concat_vars[0])):
                    _vars = ''
                    for concat_var in concat_vars:
                        _vars += '_' + str(concat_var[i])
                    _concat_vars.append(_vars)
            except IndexError:
                raise ValueError('Incorrect variable format: please make sure that all dimensions have the same number of variables')
            unique_dim_vars = []
            unique_indices = []
            for ci, _concat_var in enumerate(_concat_vars):
                if _concat_var not in unique_dim_vars:
                    unique_dim_vars.append(_concat_var)
                    unique_indices.append(ci)
            dims[idx]['size'] = len(unique_indices)
            for vars in concat_vars:
                dims[idx]['dim_vars'].append({
                    'values': [vars[ui] for ui in unique_indices]
                })
        else:
            unique_dim_vars = []
            for _dim_var in concat_vars[0]:
                if _dim_var not in unique_dim_vars:
                    unique_dim_vars.append(_dim_var)
            dims[idx]['size'] = len(unique_dim_vars)
            dims[idx]['dim_vars'].append({
                'values': unique_dim_vars
            })

    # validate interlaced dims (ensure there are no repeating columns)
    dim_ci = len(template_dim_vars[0]) + 1
    total_interlace_width = 1
    n_data_vars = len(brick_ds['dataValues'])
    for dim in dims[1:]:
         total_interlace_width *= dim['size']
    _validate_interlaced_dims(sheet,
                              n_data_vars,
                              start_row,
                              start_row + interlace_dim_var_length - 1,
                              dim_ci + 1,
                              dim_ci + total_interlace_width)

    # Get N dimensional data
    start_row = dim_start_row + 1
    start_col = len(dims[0]['dim_vars']) + 2
    len_cols = len(brick_ds['dataValues'])
    for dim in dims[1:]:
        len_cols *= dim['size']
    for di, data_var_field in enumerate(brick_ds['dataValues']):
        _data_vars = []
        for row in sheet.iter_rows(min_row=start_row,
                                   max_row=start_row + dims[0]['size'] - 1,
                                   min_col=start_col,
                                   max_col=start_col + len_cols - 1):
            dv_row = [cell.value for ci, cell in enumerate(row) if (ci + di) % len(brick_ds['dataValues']) == 0]
           # _data_vars.append(_get_nd_interlace_data(dv_row, dims[1:]))
            if len(dims) == 1:
                _data_vars.append(dv_row[0])
            else:
                _data_vars.append(_get_nd_interlace_data(dv_row, dims[1:]))
        data_vars.append({'values': _data_vars})
    _validate(dims, data_vars)
    return {
        'dims': dims,
        'data_vars': data_vars
    }

def parse_brick_tab_data(brick_ds, book):
    dims_ds = brick_ds['dimensions']
    template_dims = [dim['type']['text'] for dim in dims_ds]
    template_dim_vars = []
    for di, dim in enumerate(dims_ds):
        template_dim_vars.append([])
        for dim_var in dim['variables']:
            var_name = _get_contexted_var_name(dim_var)
            template_dim_vars[-1].append('%s:%s' % (template_dims[di], var_name))

    dims = [{'size': 0, 'dim_vars': []} for _ in dims_ds]
    data_vars = []

    #sheet = book.active # TODO: add validation to ensure dim vars are never the same across sheets
    sheet = book.worksheets[0]
    start_row = len(dims_ds) + len(brick_ds['dataValues']) + 7
    # Do dimension 1 variables
    interlace_dim_var_length = 0
    for dim in template_dim_vars[1:]:
        interlace_dim_var_length += len(dim)
    dim1_start_row = start_row + interlace_dim_var_length
    dim1_ci = 0
    for dim_var in template_dim_vars[0]:
        dim1_ci += 1
        header = sheet.cell(row=dim1_start_row, column=dim1_ci).value
        if header != dim_var:
            raise ValueError('Wrong format: expected dimension variable %s; found %s' % (dim_var, header))
        vals = _get_1d_column_data(sheet, dim1_start_row + 1, dim1_ci)
        dims[0]['size'] = len(vals)
        dims[0]['dim_vars'].append({
            'values': vals
        })

    # do interlaced dimension variables
    dim_start_row = start_row
    dim_ci = len(template_dim_vars[0]) + 1
    for di, dim_vars in enumerate(template_dim_vars[1:]):
        idx = di + 1
        concat_vars = []
        for dim_var in dim_vars:
            header = sheet.cell(row=dim_start_row, column=dim_ci).value
            if header != dim_var:
                raise ValueError('Wrong format: expected dimension variable %s, found %s' % (dim_var, header))
            raw_vars = _get_1d_row_data(sheet, dim_start_row, dim_ci + 1)
            concat_vars.append(raw_vars)
            dim_start_row += 1
        if len(concat_vars) > 1:
            _concat_vars = []
            try:
                for i in range(len(concat_vars[0])):
                    _vars = ''
                    for concat_var in concat_vars:
                        _vars += '_' + str(concat_var[i])
                    _concat_vars.append(_vars)
            except IndexError:
                raise ValueError('Incorrect variable format: please make sure that all dimensions have the same number of variables')
            unique_dim_vars = []
            unique_indices = []
            for ci, _concat_var in enumerate(_concat_vars):
                if _concat_var not in unique_dim_vars:
                    unique_dim_vars.append(_concat_var)
                    unique_indices.append(ci)
            dims[idx]['size'] = len(unique_indices)
            for vars in concat_vars:
                dims[idx]['dim_vars'].append({
                    'values': [vars[ui] for ui in unique_indices]
                })
        else:
            unique_dim_vars = []
            for _dim_var in concat_vars[0]:
                if _dim_var not in unique_dim_vars:
                    unique_dim_vars.append(_dim_var)
            dims[idx]['size'] = len(unique_dim_vars)
            dims[idx]['dim_vars'].append({
                'values': unique_dim_vars
           })

    # validate interlaced dimensions (ensure there are no repeating columns)
    dim_ci = len(template_dim_vars[0]) + 1
    total_interlace_width = 1
    for dim in dims[1:]:
        total_interlace_width *= dim['size']
    _validate_interlaced_dims(sheet,
                              1,
                              start_row,
                              start_row + interlace_dim_var_length - 1,
                              dim_ci + 1,
                              dim_ci + total_interlace_width)
    # Get N Dimensional data
    start_row = dim_start_row + 1
    start_col = len(dims[0]['dim_vars']) + 2
    len_cols = 1
    for dim in dims[1:]:
        len_cols *= dim['size']

    for _sheet in book.worksheets:
        _data_vars = []
        for row in _sheet.iter_rows(min_row=start_row,
                                   max_row=start_row + dims[0]['size'] - 1,
                                   min_col=start_col,
                                   max_col=start_col + len_cols - 1):
            dv_row = [cell.value for cell in row]
            _data_vars.append(_get_nd_interlace_data(dv_row, dims[1:]))
        data_vars.append({'values': _data_vars})

    _validate(dims, data_vars)
    return {
        'dims': dims,
        'data_vars': data_vars
    }

def parse_brick_tab_dim(brick_ds, book):
    dims_ds = brick_ds['dimensions']

    template_dims = [dim['type']['text'] for dim in dims_ds]
    template_dim_vars = []
    for di, dim in enumerate(dims_ds):
        template_dim_vars.append([])
        for dim_var in dim['variables']:
            var_name = _get_contexted_var_name(dim_var)
            template_dim_vars[-1].append('%s:%s' % (template_dims[di], var_name))

    dims = [{'size': 0, 'dim_vars': []} for _ in dims_ds]
    data_vars = []

    # TODO: make validation to ensure that all dim vars across the tabs are the same
    sheet = book.worksheets[0]

    start_row = len(dims_ds) + len(brick_ds['dataValues']) + 10

    # Do dimension 1 variables
    interlace_dim_var_length = 0
    for dim in template_dim_vars[1:-1]:
        interlace_dim_var_length += len(dim)
    dim1_start_row = start_row + interlace_dim_var_length
    dim1_ci = 0
    for dim_var in template_dim_vars[0]:
        dim1_ci += 1
        header = sheet.cell(row=dim1_start_row, column=dim1_ci).value
        if header != dim_var:
            raise ValueError('Wrong format: expected dimension variable %s; found %s' % (dim_var, header))
        vals = _get_1d_column_data(sheet, dim1_start_row + 1, dim1_ci)
        dims[0]['size'] = len(vals)
        dims[0]['dim_vars'].append({
            'values': vals
        })
    # Do interlaced dimension variables
    dim_start_row = start_row
    dim_ci = len(template_dim_vars[0]) + 1
    for di,  dim_vars in enumerate(template_dim_vars[1:-1]):
        idx = di + 1
        concat_vars = []
        for dim_var in dim_vars:
            header = sheet.cell(row=dim_start_row, column=dim_ci).value
            if header != dim_var:
                raise ValueError('Wrong format: expected dimension variable %s; found %s' % (dim_var, header))
            raw_vars = _get_1d_row_data(sheet, dim_start_row, dim_ci + 1)
            concat_vars.append(raw_vars)
            dim_start_row += 1
        if len(concat_vars) > 1:
            _concat_vars = []
            try:
                for i in range(len(concat_vars[0])):
                    _vars = ''
                    for concat_var in concat_vars:
                       _vars += '_' + str(concat_var[i])
                    _concat_vars.append(_vars)
            except IndexError:
                raise ValueError('Incorrect variable format: please make sure that all dimensions have the same number of variables')
            unique_dim_vars = []
            unique_indices = []
            for ci, _concat_var in enumerate(_concat_vars):
                if _concat_var not in unique_dim_vars:
                    unique_dim_vars.append(_concat_var)
                    unique_indices.append(ci)
            dims[idx]['size'] = len(unique_indices)
            for vars in concat_vars:
                dims[idx]['dim_vars'].append({
                    'values': [vars[ui] for ui in unique_indices]
                })
        else:
            unique_dim_vars = []
            for _dim_var in concat_vars[0]:
                if _dim_var not in unique_dim_vars:
                    unique_dim_vars.append(_dim_var)
            dims[idx]['size'] = len(unique_dim_vars)
            dims[idx]['dim_vars'].append({
                'values': unique_dim_vars
            })

    # validate interlaced dims (ensure there are no repeating columns)
    dim_ci = len(template_dim_vars[0]) + 1
    total_interlace_width = 1
    n_data_vars = len(brick_ds['dataValues'])
    for dim in dims[1:-1]:
        total_interlace_width *= dim['size']
    _validate_interlaced_dims(sheet,
                              n_data_vars,
                              start_row,
                              start_row + interlace_dim_var_length - 1,
                              dim_ci + 1,
                              dim_ci + total_interlace_width)

    # Do tabbed last dimension values

    tab_dim_vars = [[] for _ in template_dim_vars[-1]]
    tab_dim_start_row = start_row - 1
    for _sheet in book.worksheets:
        empty_vals = False
        for di, dim_var in enumerate(template_dim_vars[-1]):
            header = _sheet.cell(row=tab_dim_start_row - 1, column=di + 1).value
            if header != dim_var:
                raise ValueError('Wrong format: expected dimension variable %s, found %s' % (dim_var, header))
            value = _sheet.cell(row=tab_dim_start_row, column=di + 1).value
            if value is None:
                empty_vals = True
                break
            tab_dim_vars[di].append(value)
        if empty_vals:
            break
    tab_var_it = iter(tab_dim_vars)
    tab_len = len(next(tab_var_it))
    if not all(len(l) == tab_len for l in tab_var_it):
        raise ValueError('Incorrect variable format: please make sure that %s has the same number of dimension variables in each separate sheet'
                         % brick_ds['dimensions'][-1]['type']['text'])
    dims[-1]['size'] = tab_len
    for tab_dim_var in tab_dim_vars:
        dims[-1]['dim_vars'].append({
            'values': tab_dim_var
        })

    # Get N Dimensional data
    data_start_row = start_row + interlace_dim_var_length + 1
    data_ci = len(dims[0]['dim_vars']) + 2
    len_cols = len(brick_ds['dataValues'])
    for dim in dims[1:-1]:
        len_cols *= dim['size']

    for ri in range(len(brick_ds['dataValues'])):
        _data_vars = []
        for row in sheet.iter_rows(min_row=data_start_row,
                                   max_row=data_start_row + dims[0]['size'] - 1,
                                   min_col=data_ci,
                                   max_col=data_ci + len_cols - 1):
            dv_row = [cell.value for ci, cell in enumerate(row) if (ci + ri) % len(brick_ds['dataValues']) == 0]
            _data_vars.append(_get_nd_interlace_data(dv_row, dims[1:]))
        _data_vars_xr = xr.DataArray(_data_vars)
        for _sheet in book.worksheets[1:dims[-1]['size']]:
            _data_vars_concat = []
            for row in _sheet.iter_rows(min_row=data_start_row,
                                        max_row=data_start_row + dims[0]['size'] - 1,
                                        min_col=data_ci,
                                        max_col=data_ci + len_cols - 1):
                dv_row = [cell.value for ci, cell in enumerate(row) if (ci + ri) % len(brick_ds['dataValues']) == 0]
                _data_vars_concat.append(_get_nd_interlace_data(dv_row, dims[1:]))
            _data_vars_concat_xr = xr.DataArray(_data_vars_concat)
            _data_vars_xr = xr.concat([_data_vars_xr, _data_vars_concat_xr], dim='dim_%d' % (len(dims) - 1))
        data_vars.append({'values': _data_vars_xr.data.tolist()})

    _validate(dims, data_vars)
    return {
        'dims': dims,
        'data_vars': data_vars
    }

def _get_nd_interlace_data(row, dims):
    if len(dims) == 1:
        return row

    dim = dims[0]
    if len(row) % dim['size'] != 0:
        # TODO: make sure this is a clear enough error message
        raise ValueError('Incorrect variable format: please make sure that all dimensions have the same number of variables')

    interlace_row = []
    n_intervals = int(len(row) / dim['size'])
    for i in range(0, len(row), n_intervals):
        interlace_row.append(_get_nd_interlace_data(row[i:i + n_intervals], dims[1:]))
    return interlace_row

