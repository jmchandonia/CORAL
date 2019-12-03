import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
 

def generate_brick_F1DM_template(brick_ds,file_name):

    TITLE_FONT = Font(name='Calibri', size=14) 
    INSTRUCTIONS_FONT = Font(name='Calibri', size=10, bold=True, italic=True)
    DIM1_FILL = PatternFill('solid', fgColor="FFFFDD")  
    DATA_FILL = PatternFill('solid', fgColor="FFDDDD") 
    COLUMN_WIDTH = 18
    dim1_row_counts = 10

    book = Workbook()
    sheet = book.active

    format_name = 'F1DM'

    data_vars = brick_ds['dataValues']
    brick_dims = brick_ds['dimensions']


    template_name = brick_ds['template_type']
    data_type = brick_ds['type']['text']
    data_vars = [ _get_contexted_var_name(dvar) for dvar in data_vars ] 

    dims = [brick_dims[0]['type']['text']]
    dim1_vars = []
    for dim_var in brick_dims[0]['variables']:
        var_name = _get_contexted_var_name(dim_var)
        dim1_vars.append( '%s:%s' % (dims[0], var_name)  )

    rows = [
        ['Automatically generated template '],
        ['Name', template_name],
        ['Type', data_type],    
        ['Dimension 1', dims[0]],
        ['Data values', ', '.join(data_vars)],
        [],
        ['Instructions: fill in all colored parts of the spreadheet bellow with your data and metadata'],
        ['@Format:%s' % format_name]
    ]
        
    row = []
    for dim1_var in dim1_vars:
        row.append(dim1_var)
    for data_var in data_vars:
        row.append(data_var)
    rows.append(row)
        

    for row in rows:
        sheet.append(row)
        
    sheet['A1'].font = TITLE_FONT

    sheet['A4'].fill = DIM1_FILL
    sheet['B4'].fill = DIM1_FILL

    sheet['A5'].fill = DATA_FILL
    sheet['B5'].fill = DATA_FILL

    sheet['A7'].font = INSTRUCTIONS_FONT
    sheet['A8'].font = INSTRUCTIONS_FONT

    # Color Dim 1
    start_row = len(rows) + 1        
    start_column = 1
    for ri in range(start_row,start_row + dim1_row_counts):
        for ci,_ in enumerate(dim1_vars):
            sheet.cell(row=ri, column=ci+start_column).fill = DIM1_FILL 

        
    # Color Data    
    start_row = len(rows) + 1       
    start_column = len(dim1_vars) + 1    
    for ri in range(start_row,start_row + dim1_row_counts):
        for ci in range(start_column,start_column + len(data_vars)):
            sheet.cell(row=ri, column=ci ).fill = DATA_FILL 
        
        
    for i in range(1, len(dim1_vars) + len(data_vars) + 1 ):
        c = get_column_letter(i)
        sheet.column_dimensions[c].width = COLUMN_WIDTH   

    book.save(file_name)    


def parse_brick_F1DM_data(brick_ds, file_name):
    SEARCH_FORMAT_MAX_ROW = 100
    TEMPLATE_FORMAT = 'F1DM'
    book = load_workbook(file_name)
    sheet = book.active

    dims_ds = brick_ds['dimensions']

    template_data_vars = [ _get_contexted_var_name(dvar) for dvar in brick_ds['dataValues'] ] 
    template_dim = dims_ds[0]['type']['text']
    template_dim1_vars = []
    for dim_var in dims_ds[0]['variables']:
        var_name = _get_contexted_var_name(dim_var)
        template_dim1_vars.append( '%s:%s' % (template_dim, var_name)  )

    dims = [{
        'size' : 0,
        'dim_vars': []
    }]
    data_vars = []

    template_format = None
    start_row = -1
    for ri in range(1,SEARCH_FORMAT_MAX_ROW):
        value = sheet.cell(row=ri, column=1 ).value
        if value is not None and value.startswith('@Format:'):
            start_row = ri
            template_format = value[len('@Format:'):]
            break

    if template_format is None:
        raise ValueError('Can not identify the template format')
    
    if template_format != TEMPLATE_FORMAT:
        raise ValueError('Wrong data format: the expected foramt is %s; found %s' % (
            TEMPLATE_FORMAT, template_format))

    start_row += 1
    ci = 0

    # Do dimension variables
    for dim_var in template_dim1_vars:
        ci += 1
        header = sheet.cell(row=start_row, column=ci ).value
        if header != dim_var:
            raise ValueError('Wrong format: expected dimension variable %s; found %s' % (dim_var, header))
        vals = _get_1d_column_data(sheet, start_row + 1, ci)
        dims[0]['size'] = len(vals)
        dims[0]['dim_vars'].append({
            'values': vals            
        })

    # Do data variables
    for data_var in template_data_vars:
        ci += 1
        header = sheet.cell(row=start_row, column=ci ).value
        if header != data_var:
            raise ValueError('Wrong format: expected data variable %s; found %s' % (data_var, header))
        vals = _get_1d_column_data(sheet, start_row + 1, ci)
        data_vars.append({
            'values': vals                   
        })            

    _validate(dims, data_vars)
    return {
        'dims': dims,
        'data_vars': data_vars
    }

def generate_brick_F2D_template(brick_ds,file_name):

    TITLE_FONT = Font(name='Calibri', size=14) 
    INSTRUCTIONS_FONT = Font(name='Calibri', size=10, bold=True, italic=True)
    DIM1_FILL = PatternFill('solid', fgColor="FFFFDD")  
    DIM2_FILL = PatternFill('solid', fgColor="DDFFDD")
    DATA_FILL = PatternFill('solid', fgColor="FFDDDD") 
    COLUMN_WIDTH = 18
    dim1_row_counts = 10
    dim2_row_counts = 10

    FORMAT_NAME = 'F2D'


    book = Workbook()
    sheet = book.active

    data_vars = brick_ds['dataValues']
    brick_dims = brick_ds['dimensions']

    template_name = brick_ds['template_type']
    data_type = brick_ds['type']['text']
    data_var = _get_contexted_var_name(data_vars[0])

    dims = [brick_dims[0]['type']['text'], brick_dims[1]['type']['text']]
    dim1_vars = []
    for dim_var in brick_dims[0]['variables']:
        var_name = _get_contexted_var_name(dim_var)
        dim1_vars.append( '%s:%s' % (dims[0], var_name)  )

    dim2_vars = []
    for dim_var in brick_dims[1]['variables']:
        var_name = _get_contexted_var_name(dim_var)
        dim2_vars.append('%s:%s' % (dims[1], var_name) )

    rows = [
        ['Automatically generated template '],
        ['Name', template_name],
        ['Type', data_type],    
        ['Dimension 1', dims[0]],
        ['Dimension 2', dims[1]],
        ['Data values', data_var],
        [],
        ['Instructions: fill in all colored parts of the spreadheet bellow with your data and metadata'],
        ['@Format:%s' % FORMAT_NAME]
    ]


    for j in range(len(dim2_vars) -1 ):
        row = []
        for i in range(len(dim1_vars)):
            row.append('')
        row.append(dim2_vars[j])
        rows.append(row)
        
    row = []
    for dim1_var in dim1_vars:
        row.append(dim1_var)
    row.append(dim2_vars[-1])
    rows.append(row)
        

    for row in rows:
        sheet.append(row)
        
    sheet['A1'].font = TITLE_FONT

    sheet['A4'].fill = DIM1_FILL
    sheet['B4'].fill = DIM1_FILL

    sheet['A5'].fill = DIM2_FILL 
    sheet['B5'].fill = DIM2_FILL 

    sheet['A6'].fill = DATA_FILL
    sheet['B6'].fill = DATA_FILL

    sheet['A8'].font = INSTRUCTIONS_FONT
    sheet['A9'].font = INSTRUCTIONS_FONT

    # Color Dim 1
    start_row = len(rows) + 1        
    start_column = 1
    for ri in range(start_row,start_row + dim1_row_counts):
        for ci,_ in enumerate(dim1_vars):
            sheet.cell(row=ri, column=ci+start_column).fill = DIM1_FILL 

    # Color Dim 2        
    start_row = len(rows) - len(dim2_vars) + 1       
    start_column = len(dim1_vars) + 2
    for ri,_ in enumerate(dim2_vars):
        for ci in range(start_column,start_column + dim2_row_counts):
            sheet.cell(row=ri+ start_row, column=ci ).fill = DIM2_FILL 
        
    # Color Data    
    start_row = len(rows) + 1       
    start_column = len(dim1_vars) + 2    
    for ri in range(start_row,start_row + dim1_row_counts):
        for ci in range(start_column,start_column + dim2_row_counts):
            sheet.cell(row=ri, column=ci ).fill = DATA_FILL 
        
        
    for i in range(1, len(dim1_vars) + 2 ):
        c = get_column_letter(i)
        sheet.column_dimensions[c].width = COLUMN_WIDTH   

    book.save(file_name)    

def parse_brick_F2D_data(brick_ds, file_name):
    SEARCH_FORMAT_MAX_ROW = 100
    TEMPLATE_FORMAT = 'F2D'
    book = load_workbook(file_name)
    sheet = book.active

    dims_ds = brick_ds['dimensions']

    template_dim1 = dims_ds[0]['type']['text']
    template_dim1_vars = []
    for dim_var in dims_ds[0]['variables']:
        var_name = _get_contexted_var_name(dim_var)
        template_dim1_vars.append( '%s:%s' % (template_dim1, var_name)  )

    template_dim2 = dims_ds[1]['type']['text']
    template_dim2_vars = []
    for dim_var in dims_ds[1]['variables']:
        var_name = _get_contexted_var_name(dim_var)
        template_dim2_vars.append( '%s:%s' % (template_dim2, var_name)  )


    dims = [
        {
            'size' : 0,
            'dim_vars': []
        },{
            'size' : 0,
            'dim_vars': []
        }]
    data_vars = []

    template_format = None
    start_row = -1
    for ri in range(1,SEARCH_FORMAT_MAX_ROW):
        value = sheet.cell(row=ri, column=1 ).value
        if value is not None and value.startswith('@Format:'):
            start_row = ri
            template_format = value[len('@Format:'):]
            break

    if template_format is None:
        raise ValueError('Can not identify the template format')
    
    if template_format != TEMPLATE_FORMAT:
        raise ValueError('Wrong data format: the expected foramt is %s; found %s' % (
            TEMPLATE_FORMAT, template_format))

    start_row += 1

    # Do dimension1 variables
    dim1_start_row = start_row + len(template_dim2_vars) - 1
    dim1_ci = 0
    for dim_var in template_dim1_vars:
        dim1_ci += 1
        header = sheet.cell(row=dim1_start_row, column=dim1_ci ).value
        if header != dim_var:
            raise ValueError('Wrong format: expected dimension variable %s; found %s' % (dim_var, header))
        vals = _get_1d_column_data(sheet, dim1_start_row + 1, dim1_ci)
        dims[0]['size'] = len(vals)
        dims[0]['dim_vars'].append({
            'values': vals            
        })

    # Do dimension2 variables
    dim2_start_row = start_row
    dim2_ci = len(template_dim1_vars) + 1
    for dim_var in template_dim2_vars:
        header = sheet.cell(row=dim2_start_row, column=dim2_ci ).value
        if header != dim_var:
            raise ValueError('Wrong format: expected dimension variable %s; found %s' % (dim_var, header))
        vals = _get_1d_row_data(sheet, dim2_start_row, dim2_ci + 1)
        dims[1]['size'] = len(vals)
        dims[1]['dim_vars'].append({
            'values': vals            
        })
        dim2_start_row += 1


    # Do data variable
    ri = start_row + len(template_dim2_vars) - 1
    ci = len(template_dim1_vars) + 1 + 1
    vals = _get_2d_data(sheet, ri, ci, dims[0]['size'], dims[1]['size'])
    data_vars.append({
        'values': vals                   
    })            

    _validate(dims, data_vars)
    return {
        'dims': dims,
        'data_vars': data_vars
    }


def parse_brick_F2D_CSV_data(brick_ds, file_name):
    # Read df
    df = pd.read_csv(file_name, sep='\t') 

    # Build vairable values for each dims
    dims = []

    # First dim
    brick_dims = brick_ds['dimensions'] 
    if len(brick_dims) > 0:
        dim = {
            'size':  0,
            'dim_vars': []
        }
        dims.append(dim)
        for i,v in enumerate(brick_dims[0]['variables']):
            vals = list(df[df.columns[i]].values)
            dim['size'] = len(vals)
            dim['dim_vars'].append({
                'values': vals
            })

    # Second dim
    if len(brick_dims) > 1:
        dim = {
            'size':  0,
            'dim_vars': []
        }
        dims.append(dim)
        offset = len(brick_dims[0]['variables'])
        vals = list(df.columns[offset:].values)
        dim['size'] = len(vals)
        dim['dim_vars'].append({
                'values': vals
            })

    data_vars = []
    offset = len(dims[0]['dim_vars'])
    vals = df[df.columns[offset:].values].values
    data_vars.append({
                'values': vals
            })

    _validate(dims, data_vars)    
    return {
        'dims': dims,
        'data_vars': data_vars
    }

def _validate(dims, data_vars):  
    dim_sizes = []

    # Validate dimension variables
    for dim in dims:
        dim_size = None
        for dim_var in dim['dim_vars']:
            size = len(dim_var['values'])
            if dim_size is None:
                dim_size = size
            else:
                if dim_size != size:
                    raise ValueError('Size of two vars from the same dimension is different: %s and %s' % (dim_size, size))  
        dim_sizes.append(dim_size)

    # Validate data vars
    for data_var in data_vars:
        data_shape = list(np.array(data_var['values']).shape)

        if len(data_shape) != len(dim_sizes):
            raise ValueError('The dimensionality of data (%s) is different from the context (%s)' % (len(data_shape), len(dim_sizes)))

        for dim_index, dim_size in enumerate(dim_sizes):
            if dim_size != data_shape[dim_index]:
                raise ValueError('The size of dimension for data var (%s) and context (%s) ' % (dim_size, data_shape[dim_index]))

def _get_1d_column_data(sheet, ri, ci):
    data = []
    while True:
        value = sheet.cell(row=ri, column=ci ).value
        if value is None:
            break
        data.append(value)
        ri += 1
    return data

def _get_1d_row_data(sheet, ri, ci):
    data = []
    while True:
        value = sheet.cell(row=ri, column=ci ).value
        if value is None:
            break
        data.append(value)
        ci += 1
    return data

def _get_2d_data(sheet, ri_start, ci_start, r_len, c_len):
    data = []
    for ri in range(ri_start, ri_start + r_len):
        row = []
        for ci in range(ci_start, ci_start + c_len):
            value = sheet.cell(row=ri, column=ci).value
            row.append(value)
        data.append(row)
    return data


def _get_contexted_var_name(dvar):
    var_name = dvar['type']['text']
    if 'units' in dvar and dvar['units'] and dvar['units']['text']:
        var_name += ' (%s)' % dvar['units']['text']

    items = []
    if 'context' in dvar:
        for ce in dvar['context']:
            item = '%s=%s' % (ce['type']['text'], ce['value']['text'])
            if 'units' in ce and ce['units'] and ce['units']['text']:
                item += ' (%s)' % ce['units']['text']
            items.append(item)

    return var_name if len(items) == 0 else '%s:%s' % (var_name, '; '.join(items))
